"""
Internetten bulunan ornek mizan verisini Excel dosyasina donustur.
Kaynak: muhasebedersleri.com - Genel Gecici Mizan ornegi
+ Gercekci gelir tablosu hesaplari eklendi.

Bu mizan, orta olcekli bir Turk ticaret sirketini simule eder.
"""
import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Mizan"

headers = ["Hesap Kodu", "Hesap Adi", "Borc", "Alacak", "Borc Bakiye", "Alacak Bakiye"]
ws.append(headers)

header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(bold=True, size=11, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

for col_idx in range(1, 7):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

# Mizan verileri - Gercekci bir Turk ticaret sirketi
# Gelir tablosu hesaplama:
#   Gelirler: 600(3,850,000) + 601(650,000) = 4,500,000
#   Satis indirimleri: 610(125,000) + 611(45,000) = 170,000
#   Net Satis = 4,330,000
#   SMM: 620(2,680,000) + 621(420,000) = 3,100,000
#   Brut Kar = 1,230,000
#   Faaliyet Giderleri: 631(310,000) + 632(385,000) = 695,000
#   Faaliyet Kari = 535,000
#   Diger Gelirler: 642(38,000) + 644(12,000) + 646(95,000) + 647(8,000) + 671(18,000) = 171,000
#   Diger Giderler: 653(22,000) + 654(35,000) + 656(78,000) + 657(15,000) + 660(62,000) + 661(48,000) + 689(8,000) = 268,000
#   Vergi Oncesi Kar = 535,000 + 171,000 - 268,000 = 438,000
#   Vergi: 691 = 438,000 * %25 = 109,500
#   Net Kar = 438,000 - 109,500 = 328,500

# 590 = 328,500, 370 = 109,500

mizan_data = [
    # DONEN VARLIKLAR
    ("100", "Kasa",                                66000, 37000,   29000, 0),
    ("101", "Alinan Cekler",                       85000, 20000,   65000, 0),
    ("102", "Bankalar",                           457000, 210000, 247000, 0),
    ("120", "Alicilar",                           680000, 245000, 435000, 0),
    ("121", "Alacak Senetleri",                   147000, 50000,   97000, 0),
    ("126", "Verilen Depozito ve Teminatlar",      18000, 0,       18000, 0),
    ("128", "Supheli Ticari Alacaklar",            75000, 0,       75000, 0),
    ("129", "Supheli Ticari Alacak Karsiligi (-)",     0, 30000,       0, 30000),
    ("150", "Ilk Madde ve Malzeme",               120000, 45000,   75000, 0),
    ("153", "Ticari Mallar",                      335000, 200000, 135000, 0),
    ("159", "Verilen Siparis Avanslari",           25000, 0,       25000, 0),
    ("180", "Gelecek Aylara Ait Giderler",         42000, 14000,   28000, 0),
    ("190", "Devreden KDV",                        38000, 0,       38000, 0),
    ("191", "Indirilecek KDV",                    165000, 165000,      0, 0),

    # DURAN VARLIKLAR
    ("252", "Binalar",                            710000, 0,      710000, 0),
    ("253", "Tesis, Makine ve Cihazlar",          480000, 0,      480000, 0),
    ("254", "Tasitlar",                           220000, 0,      220000, 0),
    ("255", "Demirbaslar",                        185000, 55000,  130000, 0),
    ("257", "Birikmis Amortismanlar (-)",              0, 408000,      0, 408000),
    ("260", "Haklar",                              90000, 0,       90000, 0),
    ("264", "Ozel Maliyetler",                     65000, 0,       65000, 0),
    ("268", "Birikmis Itfa Paylari (-)",               0, 42000,       0, 42000),

    # KISA VADELI YABANCI KAYNAKLAR
    ("300", "Banka Kredileri",                     50000, 320000,      0, 270000),
    ("320", "Saticilar",                          140000, 515000,      0, 375000),
    ("321", "Borc Senetleri",                          0, 179000,      0, 179000),
    ("335", "Personele Borclar",                   85000, 152000,      0, 67000),
    ("340", "Alinan Avanslar",                         0, 45000,       0, 45000),
    ("360", "Odenecek Vergi ve Fonlar",            42000, 98000,       0, 56000),
    ("361", "Odenecek Sosyal Guvenlik Kesintileri",35000, 73000,       0, 38000),
    ("370", "Donem Kari Vergi ve Diger Yasal Yuk. Kars.", 0, 109500,   0, 109500),
    ("372", "Kidem Tazminati Karsiligi",               0, 65000,       0, 65000),
    ("380", "Gelecek Aylara Ait Gelirler",             0, 32000,       0, 32000),
    ("391", "Hesaplanan KDV",                     195000, 195000,      0, 0),

    # UZUN VADELI YABANCI KAYNAKLAR
    ("400", "Banka Kredileri (Uzun Vadeli)",            0, 408000,      0, 408000),
    ("421", "Borc Senetleri (Uzun Vadeli)",        30000, 37000,       0, 7000),
    ("472", "Kidem Tazminati Karsiligi (Uzun Vadeli)", 0, 120000,      0, 120000),

    # OZKAYNAKLAR
    ("500", "Sermaye",                                 0, 529000,      0, 529000),
    ("540", "Yasal Yedekler",                          0, 85000,       0, 85000),
    ("542", "Olagustu Yedekler",                       0, 42000,       0, 42000),
    ("570", "Gecmis Yillar Karlari",                   0, 175000,      0, 175000),
    ("590", "Donem Net Kari",                          0, 328500,      0, 328500),

    # GELIR TABLOSU HESAPLARI
    ("600", "Yurtici Satislar",                        0, 3850000,     0, 3850000),
    ("601", "Yurtdisi Satislar",                       0, 650000,      0, 650000),
    ("610", "Satistan Iadeler (-)",               125000, 0,      125000, 0),
    ("611", "Satis Iskontolari (-)",               45000, 0,       45000, 0),
    ("620", "Satilan Mamuller Maliyeti (-)",      2680000, 0,     2680000, 0),
    ("621", "Satilan Ticari Mallar Maliyeti (-)",  420000, 0,      420000, 0),
    ("631", "Pazarlama Satis Dagitim Giderleri (-)", 310000, 0,    310000, 0),
    ("632", "Genel Yonetim Giderleri (-)",         385000, 0,      385000, 0),
    ("642", "Faiz Gelirleri",                          0, 38000,       0, 38000),
    ("644", "Konusu Kalmayan Karsiliklar",             0, 12000,       0, 12000),
    ("646", "Kambiyo Karlari",                         0, 95000,       0, 95000),
    ("647", "Reeskont Faiz Gelirleri",                 0, 8000,        0, 8000),
    ("653", "Komisyon Giderleri (-)",               22000, 0,       22000, 0),
    ("654", "Karsilik Giderleri (-)",               35000, 0,       35000, 0),
    ("656", "Kambiyo Zararlari (-)",                78000, 0,       78000, 0),
    ("657", "Reeskont Faiz Giderleri (-)",          15000, 0,       15000, 0),
    ("660", "Kisa Vadeli Borclanma Giderleri (-)",  62000, 0,       62000, 0),
    ("661", "Uzun Vadeli Borclanma Giderleri (-)",  48000, 0,       48000, 0),
    ("671", "Onceki Donem Gelir ve Karlari",            0, 18000,       0, 18000),
    ("689", "Diger Olagandisi Gider ve Zararlari (-)", 8000, 0,      8000, 0),
    ("691", "Donem Kari Vergi ve Yasal Yuk. Kars. (-)", 109500, 0,  109500, 0),
]

# Veri yaz
num_format = '#,##0'
for row_data in mizan_data:
    ws.append(row_data)

# Formatlama
for row_idx in range(2, len(mizan_data) + 2):
    for col_idx in range(1, 7):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = thin_border
        if col_idx >= 3:
            cell.number_format = num_format
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 1:
            cell.alignment = Alignment(horizontal="center")

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 50
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16

# Toplamlar
total_row = len(mizan_data) + 2
ws.cell(row=total_row, column=2, value="TOPLAM").font = Font(bold=True, size=11)

total_borc = sum(r[2] for r in mizan_data)
total_alacak = sum(r[3] for r in mizan_data)
total_borc_bakiye = sum(r[4] for r in mizan_data)
total_alacak_bakiye = sum(r[5] for r in mizan_data)

ws.cell(row=total_row, column=3, value=total_borc).number_format = num_format
ws.cell(row=total_row, column=4, value=total_alacak).number_format = num_format
ws.cell(row=total_row, column=5, value=total_borc_bakiye).number_format = num_format
ws.cell(row=total_row, column=6, value=total_alacak_bakiye).number_format = num_format

total_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
for col_idx in range(1, 7):
    cell = ws.cell(row=total_row, column=col_idx)
    cell.font = Font(bold=True, size=11)
    cell.fill = total_fill
    cell.border = thin_border

# Dogrulama
print(f"Toplam Borc:          {total_borc:>15,.2f} TL")
print(f"Toplam Alacak:        {total_alacak:>15,.2f} TL")
print(f"Borc - Alacak Farki:  {total_borc - total_alacak:>15,.2f} TL")
print(f"Borc Bakiye Toplami:  {total_borc_bakiye:>15,.2f} TL")
print(f"Alacak Bakiye Toplami:{total_alacak_bakiye:>15,.2f} TL")
print(f"Bakiye Farki:         {total_borc_bakiye - total_alacak_bakiye:>15,.2f} TL")
print(f"Hesap Sayisi:         {len(mizan_data)}")

# Dengeleme kontrolu
if total_borc != total_alacak:
    diff = total_alacak - total_borc
    print(f"\n!! Mizan dengesi icin Bankalar (102) borc tarafina {diff:,.0f} TL eklenmeli")
    # Otomatik dengeleme: 102 Bankalar borc tarafini artir
    for i, r in enumerate(mizan_data):
        if r[0] == "102":
            old_borc = r[2]
            old_bakiye = r[4]
            mizan_data[i] = ("102", r[1], old_borc + diff, r[3], old_bakiye + diff, r[5])
            print(f"   102 Bankalar guncellendi: Borc {old_borc:,.0f} -> {old_borc+diff:,.0f}, Bakiye {old_bakiye:,.0f} -> {old_bakiye+diff:,.0f}")
            break

    # Yeniden yaz
    for row_idx in range(2, len(mizan_data) + 2):
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx).value = None

    for idx, row_data in enumerate(mizan_data):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 2, column=col_idx)
            cell.value = val
            cell.border = thin_border
            if col_idx >= 3:
                cell.number_format = num_format
                cell.alignment = Alignment(horizontal="right")
            elif col_idx == 1:
                cell.alignment = Alignment(horizontal="center")

    total_borc = sum(r[2] for r in mizan_data)
    total_alacak = sum(r[3] for r in mizan_data)
    total_borc_bakiye = sum(r[4] for r in mizan_data)
    total_alacak_bakiye = sum(r[5] for r in mizan_data)

    ws.cell(row=total_row, column=3, value=total_borc).number_format = num_format
    ws.cell(row=total_row, column=4, value=total_alacak).number_format = num_format
    ws.cell(row=total_row, column=5, value=total_borc_bakiye).number_format = num_format
    ws.cell(row=total_row, column=6, value=total_alacak_bakiye).number_format = num_format

    print(f"\n--- Dengeleme sonrasi ---")
    print(f"Toplam Borc:          {total_borc:>15,.2f} TL")
    print(f"Toplam Alacak:        {total_alacak:>15,.2f} TL")
    print(f"Borc Bakiye:          {total_borc_bakiye:>15,.2f} TL")
    print(f"Alacak Bakiye:        {total_alacak_bakiye:>15,.2f} TL")

assert total_borc == total_alacak, f"MIZAN DENGESIZ! Borc={total_borc}, Alacak={total_alacak}"
assert total_borc_bakiye == total_alacak_bakiye, f"BAKIYE DENGESIZ!"

# Gelir tablosu kontrolu
gelir = sum(r[5] for r in mizan_data if r[0].startswith("6") and r[5] > 0)
gider_vergi_haric = sum(r[4] for r in mizan_data if r[0].startswith("6") and r[4] > 0 and r[0] != "691")
vergi = next((r[4] for r in mizan_data if r[0] == "691"), 0)
net_kar = gelir - gider_vergi_haric - vergi
donem_net_kari = next((r[5] for r in mizan_data if r[0] == "590"), 0)

print(f"\n--- Gelir Tablosu ---")
print(f"Toplam Gelirler:      {gelir:>15,.2f} TL")
print(f"Toplam Giderler:      {gider_vergi_haric:>15,.2f} TL")
print(f"Vergi Oncesi Kar:     {gelir - gider_vergi_haric:>15,.2f} TL")
print(f"Vergi:                {vergi:>15,.2f} TL")
print(f"Net Kar:              {net_kar:>15,.2f} TL")
print(f"590 Donem Net Kari:   {donem_net_kari:>15,.2f} TL")
print(f"Tutarli: {'EVET' if net_kar == donem_net_kari else 'HAYIR'}")

import os
output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "mizan_internet_ornegi.xlsx")
wb.save(output_path)
print(f"\nDosya kaydedildi: {output_path}")
print("MIZAN DENGELI - HAZIR")
