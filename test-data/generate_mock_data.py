"""
TFRS Converter - Mock Mizan Verisi Uretici

Bu script test amacli gercekci mizan verileri uretir:
- mizan_basit.xlsx   : ~38 hesap, temiz veri
- mizan_kapsamli.xlsx: 65+ hesap, tum TDHP gruplari
- mizan_basit.csv    : Basit mizanin CSV versiyonu

Tum mizanlarda borc toplami = alacak toplami (denge saglanir).
Gelir tablosu net kari = 590 Donem Net Kari (tutarlilik saglanir).
"""

import csv
import os
from decimal import Decimal, ROUND_HALF_UP
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def d(val):
    """Helper: float/int -> Decimal (2 ondalik)"""
    return Decimal(str(val)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def hesaplari_dengele(hesaplar):
    """
    Mizan dengeleme:
    1. Gelir tablosu net karini hesapla (6xx - 69x)
    2. 590'i bu degere ayarla
    3. Genel toplam dengesini kontrol et, fark varsa 102 Bankalar'i ayarla
    Sonuc: list of (kod, ad, borc, alacak) tuple'lari (duzeltilmis)
    """
    hesap_dict = {h[0]: list(h) for h in hesaplar}

    # Gelir tablosu net kari: (gelir alacak - gelir borc) 6xx, 69x haric
    gelir_net = sum(
        d(h[3]) - d(h[2])
        for h in hesaplar
        if h[0].startswith("6") and not h[0].startswith("69")
    )
    # Vergi: 69x (borc tarafinda)
    vergi = sum(
        d(h[2]) - d(h[3])
        for h in hesaplar
        if h[0].startswith("69")
    )
    net_kar = gelir_net - vergi

    # 590 Donem Net Kari'ni ayarla
    if "590" in hesap_dict:
        hesap_dict["590"][2] = 0  # borc
        hesap_dict["590"][3] = float(net_kar)  # alacak

    # Simdi genel dengeyi kontrol et
    result = [tuple(hesap_dict[h[0]]) for h in hesaplar]
    toplam_borc = sum(d(h[2]) for h in result)
    toplam_alacak = sum(d(h[3]) for h in result)
    fark = toplam_borc - toplam_alacak

    # Fark varsa 102 Bankalar'dan duzelt
    if fark != 0 and "102" in hesap_dict:
        eski_banka = d(hesap_dict["102"][2])
        hesap_dict["102"][2] = float(eski_banka - fark)
        result = [tuple(hesap_dict[h[0]]) for h in hesaplar]

    return result


# ---------------------------------------------------------------------------
# BASIT MIZAN VERILERI (~38 hesap)
# Gercekci bir orta olcekli Turk ticaret sirketi: ABC Ticaret A.S.
# ---------------------------------------------------------------------------
_BASIT_HESAPLAR_RAW = [
    # (hesap_kodu, hesap_adi, borc_bakiye, alacak_bakiye)
    # --- 1. Donen Varliklar ---
    ("100", "Kasa", 25_000, 0),
    ("102", "Bankalar", 1_450_000, 0),
    ("120", "Alicilar", 890_000, 0),
    ("128", "Supheli Ticari Alacaklar", 120_000, 0),
    ("129", "Supheli Ticari Alacak Karsiligi (-)", 0, 45_000),
    ("153", "Ticari Mallar", 670_000, 0),
    ("180", "Gelecek Aylara Ait Giderler", 36_000, 0),
    # --- 2. Duran Varliklar ---
    ("252", "Binalar", 3_200_000, 0),
    ("253", "Tesis, Makine ve Cihazlar", 1_850_000, 0),
    ("254", "Tasitlar", 480_000, 0),
    ("255", "Demirbaslar", 210_000, 0),
    ("257", "Birikmis Amortismanlar (-)", 0, 1_420_000),
    ("260", "Haklar", 150_000, 0),
    ("268", "Birikmis Itfa Paylari (-)", 0, 60_000),
    # --- 3. Kisa Vadeli Yabanci Kaynaklar ---
    ("300", "Banka Kredileri", 0, 750_000),
    ("320", "Saticilar", 0, 620_000),
    ("335", "Personele Borclar", 0, 185_000),
    ("360", "Odenecek Vergi ve Fonlar", 0, 95_000),
    ("361", "Odenecek Sosyal Guvenlik Kesintileri", 0, 72_000),
    ("370", "Donem Kari Vergi ve Diger Yasal Yuk. Kars.", 0, 280_000),
    ("372", "Kidem Tazminati Karsiligi", 0, 145_000),
    # --- 4. Uzun Vadeli Yabanci Kaynaklar ---
    ("400", "Banka Kredileri (Uzun Vadeli)", 0, 1_200_000),
    ("472", "Kidem Tazminati Karsiligi (Uzun Vadeli)", 0, 320_000),
    # --- 5. Ozkaynaklar ---
    ("500", "Sermaye", 0, 2_000_000),
    ("540", "Yasal Yedekler", 0, 180_000),
    ("570", "Gecmis Yillar Karlari", 0, 450_000),
    ("590", "Donem Net Kari", 0, 0),  # otomatik hesaplanacak
    # --- 6. Gelirler ---
    ("600", "Yurtici Satislar", 0, 8_500_000),
    ("601", "Yurtdisi Satislar", 0, 1_200_000),
    # Gelir Indirimleri
    ("610", "Satistan Iadeler (-)", 350_000, 0),
    # Satis Maliyeti
    ("620", "Satilan Mamul Maliyeti", 5_800_000, 0),
    # Faaliyet Giderleri
    ("631", "Pazarlama Satis Dagitim Giderleri", 680_000, 0),
    ("632", "Genel Yonetim Giderleri", 920_000, 0),
    # Diger Gelirler
    ("642", "Faiz Gelirleri", 0, 85_000),
    ("646", "Kambiyo Karlari", 0, 120_000),
    # Diger Giderler
    ("656", "Kambiyo Zararlari", 195_000, 0),
    ("660", "Kisa Vadeli Borclanma Giderleri", 290_000, 0),
    # Vergi
    ("691", "Donem Kari Vergi Karsiligi", 280_000, 0),
]

# ---------------------------------------------------------------------------
# KAPSAMLI MIZAN VERILERI (65+ hesap, tum TDHP gruplari)
# Buyuk olcekli uretim sirketi: XYZ Sanayi ve Ticaret A.S.
# ---------------------------------------------------------------------------
_KAPSAMLI_HESAPLAR_RAW = [
    # --- 1. DONEN VARLIKLAR ---
    # 10 - Hazir Degerler
    ("100", "Kasa", 18_500, 0),
    ("101", "Alinan Cekler", 245_000, 0),
    ("102", "Bankalar", 2_890_000, 0),
    ("103", "Verilen Cekler ve Odeme Emirleri (-)", 0, 165_000),
    ("108", "Diger Hazir Degerler", 12_000, 0),
    # 11 - Menkul Kiymetler
    ("110", "Hisse Senetleri", 350_000, 0),
    ("111", "Ozel Kesim Tahvil Senet ve Bonolari", 200_000, 0),
    # 12 - Ticari Alacaklar
    ("120", "Alicilar", 1_420_000, 0),
    ("121", "Alacak Senetleri", 680_000, 0),
    ("126", "Verilen Depozito ve Teminatlar", 95_000, 0),
    ("128", "Supheli Ticari Alacaklar", 280_000, 0),
    ("129", "Supheli Ticari Alacak Karsiligi (-)", 0, 110_000),
    # 13 - Diger Alacaklar
    ("131", "Ortaklardan Alacaklar", 50_000, 0),
    ("136", "Diger Cesitli Alacaklar", 38_000, 0),
    # 15 - Stoklar
    ("150", "Ilk Madde ve Malzeme", 890_000, 0),
    ("151", "Yari Mamuller - Uretim", 420_000, 0),
    ("152", "Mamuller", 1_150_000, 0),
    ("153", "Ticari Mallar", 780_000, 0),
    ("157", "Diger Stoklar", 65_000, 0),
    ("158", "Stok Deger Dusuklugi Karsiligi (-)", 0, 85_000),
    # 17 - Yillara Yaygin Insaat
    ("170", "Yillara Yaygin Insaat ve Onarim Maliyetleri", 1_500_000, 0),
    ("350", "Yillara Yaygin Insaat ve Onarim Hakedis Bedelleri", 0, 1_200_000),
    # 18 - Gelecek Aylara Ait Giderler
    ("180", "Gelecek Aylara Ait Giderler", 48_000, 0),
    ("181", "Gelir Tahakkuklari", 25_000, 0),
    # 19 - Diger Donen Varliklar
    ("190", "Devreden KDV", 320_000, 0),
    ("193", "Pesin Odenen Vergiler ve Fonlar", 180_000, 0),
    # --- 2. DURAN VARLIKLAR ---
    # 22 - Ticari Alacaklar (Uzun Vadeli)
    ("220", "Alicilar (Uzun Vadeli)", 450_000, 0),
    # 24 - Mali Duran Varliklar
    ("240", "Bagli Menkul Kiymetler", 600_000, 0),
    ("242", "Istirakler", 1_200_000, 0),
    # 25 - Maddi Duran Varliklar
    ("250", "Arazi ve Arsalar", 4_500_000, 0),
    ("251", "Yeralti ve Yerustu Duzenleri", 850_000, 0),
    ("252", "Binalar", 6_800_000, 0),
    ("253", "Tesis, Makine ve Cihazlar", 3_950_000, 0),
    ("254", "Tasitlar", 1_280_000, 0),
    ("255", "Demirbaslar", 520_000, 0),
    ("257", "Birikmis Amortismanlar (-)", 0, 4_850_000),
    ("258", "Yapilmakta Olan Yatirimlar", 750_000, 0),
    # 26 - Maddi Olmayan Duran Varliklar
    ("260", "Haklar", 380_000, 0),
    ("263", "Arastirma ve Gelistirme Giderleri", 220_000, 0),
    ("264", "Ozel Maliyetler", 160_000, 0),
    ("268", "Birikmis Itfa Paylari (-)", 0, 285_000),
    # 28 - Gelecek Yillara Ait Giderler
    ("280", "Gelecek Yillara Ait Giderler", 120_000, 0),
    ("281", "Gelir Tahakkuklari (Uzun Vadeli)", 45_000, 0),
    # --- 3. KISA VADELI YABANCI KAYNAKLAR ---
    ("300", "Banka Kredileri", 0, 2_350_000),
    ("320", "Saticilar", 0, 1_680_000),
    ("321", "Borc Senetleri", 0, 890_000),
    ("326", "Alinan Depozito ve Teminatlar", 0, 125_000),
    ("335", "Personele Borclar", 0, 420_000),
    ("360", "Odenecek Vergi ve Fonlar", 0, 285_000),
    ("361", "Odenecek Sosyal Guvenlik Kesintileri", 0, 195_000),
    ("370", "Donem Kari Vergi ve Diger Yasal Yuk. Kars.", 0, 520_000),
    ("372", "Kidem Tazminati Karsiligi", 0, 310_000),
    ("380", "Gelecek Aylara Ait Gelirler", 0, 180_000),
    ("381", "Gider Tahakkuklari", 0, 95_000),
    # --- 4. UZUN VADELI YABANCI KAYNAKLAR ---
    ("400", "Banka Kredileri (Uzun Vadeli)", 0, 3_500_000),
    ("420", "Saticilar (Uzun Vadeli)", 0, 650_000),
    ("472", "Kidem Tazminati Karsiligi (Uzun Vadeli)", 0, 580_000),
    ("480", "Gelecek Yillara Ait Gelirler", 0, 250_000),
    # --- 5. OZKAYNAKLAR ---
    ("500", "Sermaye", 0, 5_000_000),
    ("520", "Hisse Senetleri Ihrac Primleri", 0, 200_000),
    ("540", "Yasal Yedekler", 0, 480_000),
    ("542", "Olaganustu Yedekler", 0, 350_000),
    ("549", "Ozel Fonlar", 0, 120_000),
    ("570", "Gecmis Yillar Karlari", 0, 1_250_000),
    ("590", "Donem Net Kari", 0, 0),  # otomatik hesaplanacak
    # --- 6. GELIR TABLOSU ---
    # Gelirler
    ("600", "Yurtici Satislar", 0, 22_500_000),
    ("601", "Yurtdisi Satislar", 0, 3_800_000),
    ("602", "Diger Gelirler", 0, 450_000),
    # Gelir Indirimleri
    ("610", "Satistan Iadeler (-)", 680_000, 0),
    ("611", "Satis Iskontalari (-)", 320_000, 0),
    # SMM
    ("620", "Satilan Mamul Maliyeti", 12_800_000, 0),
    ("621", "Satilan Ticari Mal Maliyeti", 3_500_000, 0),
    ("622", "Satilan Hizmet Maliyeti", 850_000, 0),
    # Faaliyet Giderleri
    ("630", "Arastirma ve Gelistirme Giderleri", 420_000, 0),
    ("631", "Pazarlama Satis Dagitim Giderleri", 1_580_000, 0),
    ("632", "Genel Yonetim Giderleri", 2_150_000, 0),
    # Diger Gelirler/Giderler
    ("640", "Istiraklerden Temettue Gelirleri", 0, 180_000),
    ("642", "Faiz Gelirleri", 0, 290_000),
    ("644", "Konusu Kalmayan Karsiliklar", 0, 85_000),
    ("646", "Kambiyo Karlari", 0, 380_000),
    ("647", "Reeskont Faiz Gelirleri", 0, 55_000),
    ("649", "Diger Olagan Gelir ve Karlar", 0, 120_000),
    ("653", "Komisyon Giderleri", 95_000, 0),
    ("654", "Karsilik Giderleri", 175_000, 0),
    ("655", "Menkul Kiymet Satis Zararlari", 40_000, 0),
    ("656", "Kambiyo Zararlari", 480_000, 0),
    ("657", "Reeskont Faiz Giderleri", 65_000, 0),
    ("659", "Diger Olagan Gider ve Zararlar", 35_000, 0),
    # Olagan Disi Gelirler/Giderler
    ("671", "Onceki Donem Gelir ve Karlari", 0, 150_000),
    ("679", "Diger Olagan Disi Gelir ve Karlar", 0, 45_000),
    ("681", "Onceki Donem Gider ve Zararlari", 85_000, 0),
    # Borclanma Giderleri
    ("660", "Kisa Vadeli Borclanma Giderleri", 580_000, 0),
    ("661", "Uzun Vadeli Borclanma Giderleri", 320_000, 0),
    # Vergi
    ("691", "Donem Kari Vergi Karsiligi (-)", 520_000, 0),
]

# Dengeli versiyonlari hesapla
BASIT_HESAPLAR = hesaplari_dengele(_BASIT_HESAPLAR_RAW)
KAPSAMLI_HESAPLAR = hesaplari_dengele(_KAPSAMLI_HESAPLAR_RAW)


def hesap_denge_kontrol(hesaplar, mizan_adi):
    """Mizan dengesini kontrol et: toplam borc = toplam alacak"""
    toplam_borc = sum(d(h[2]) for h in hesaplar)
    toplam_alacak = sum(d(h[3]) for h in hesaplar)
    denge = toplam_borc == toplam_alacak
    print(f"\n{'='*60}")
    print(f"  {mizan_adi}")
    print(f"{'='*60}")
    print(f"  Toplam Borc   : {toplam_borc:>20,.2f} TL")
    print(f"  Toplam Alacak : {toplam_alacak:>20,.2f} TL")
    print(f"  Fark          : {toplam_borc - toplam_alacak:>20,.2f} TL")
    print(f"  Denge         : {'TAMAM' if denge else '*** DENGESIZ ***'}")
    print(f"  Hesap Sayisi  : {len(hesaplar)}")
    print(f"{'='*60}")
    return denge, toplam_borc, toplam_alacak


def gelir_tablosu_kontrol(hesaplar, mizan_adi):
    """Gelir tablosu net karini hesapla ve 590 ile karsilastir"""
    gelirler = sum(
        d(h[3]) - d(h[2])
        for h in hesaplar
        if h[0].startswith("6") and not h[0].startswith("69")
    )
    vergi = sum(
        d(h[2]) - d(h[3])
        for h in hesaplar
        if h[0].startswith("69")
    )
    net_kar = gelirler - vergi
    donem_net_kar = d(0)
    for h in hesaplar:
        if h[0] == "590":
            donem_net_kar = d(h[3]) - d(h[2])
            break
    eslesme = net_kar == donem_net_kar
    print(f"\n  Gelir Tablosu Kontrolu ({mizan_adi}):")
    print(f"  Gelirler - Giderler (vergi haric): {gelirler:>15,.2f} TL")
    print(f"  Vergi Karsiligi                  : {vergi:>15,.2f} TL")
    print(f"  Hesaplanan Net Kar               : {net_kar:>15,.2f} TL")
    print(f"  590 - Donem Net Kari             : {donem_net_kar:>15,.2f} TL")
    print(f"  Eslesme                          : {'TAMAM' if eslesme else '*** UYUMSUZ ***'}")
    return eslesme


def xlsx_olustur(dosya_yolu, hesaplar, sirket_adi, donem):
    """Excel mizan dosyasi olustur"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mizan"

    # Stiller
    baslik_font = Font(name="Calibri", size=14, bold=True)
    alt_baslik_font = Font(name="Calibri", size=11, bold=False, italic=True)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    grup_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    grup_font = Font(name="Calibri", size=11, bold=True)
    toplam_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    toplam_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # Baslik
    ws.merge_cells("A1:D1")
    ws["A1"] = sirket_adi
    ws["A1"].font = baslik_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:D2")
    ws["A2"] = f"Mizan Raporu - {donem}"
    ws["A2"].font = alt_baslik_font
    ws["A2"].alignment = Alignment(horizontal="center")

    # Bos satir
    row = 4

    # Header
    headers = ["Hesap Kodu", "Hesap Adi", "Borc Bakiye", "Alacak Bakiye"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    row += 1

    # TDHP Grup adlari
    grup_adlari = {
        "1": "1 - DONEN VARLIKLAR",
        "2": "2 - DURAN VARLIKLAR",
        "3": "3 - KISA VADELI YABANCI KAYNAKLAR",
        "4": "4 - UZUN VADELI YABANCI KAYNAKLAR",
        "5": "5 - OZKAYNAKLAR",
        "6": "6 - GELIR TABLOSU HESAPLARI",
    }

    onceki_grup = ""
    toplam_borc = d(0)
    toplam_alacak = d(0)

    for hesap_kodu, hesap_adi, borc, alacak in hesaplar:
        # Grup basligi ekle
        mevcut_grup = hesap_kodu[0]
        if mevcut_grup != onceki_grup and mevcut_grup in grup_adlari:
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.fill = grup_fill
                cell.border = border
            ws.cell(row=row, column=1, value=grup_adlari[mevcut_grup]).font = grup_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            ws.cell(row=row, column=1).fill = grup_fill
            ws.cell(row=row, column=1).border = border
            row += 1
            onceki_grup = mevcut_grup

        # Hesap satiri
        ws.cell(row=row, column=1, value=hesap_kodu).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=hesap_adi).border = border
        borc_cell = ws.cell(row=row, column=3, value=float(d(borc)))
        borc_cell.number_format = '#,##0.00'
        borc_cell.border = border
        borc_cell.alignment = Alignment(horizontal="right")
        alacak_cell = ws.cell(row=row, column=4, value=float(d(alacak)))
        alacak_cell.number_format = '#,##0.00'
        alacak_cell.border = border
        alacak_cell.alignment = Alignment(horizontal="right")

        toplam_borc += d(borc)
        toplam_alacak += d(alacak)
        row += 1

    # Toplam satiri
    row += 1
    for col in range(1, 5):
        cell = ws.cell(row=row, column=col)
        cell.fill = toplam_fill
        cell.font = toplam_font
        cell.border = border
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="GENEL TOPLAM").alignment = Alignment(horizontal="center")
    borc_t = ws.cell(row=row, column=3, value=float(toplam_borc))
    borc_t.number_format = '#,##0.00'
    borc_t.alignment = Alignment(horizontal="right")
    alacak_t = ws.cell(row=row, column=4, value=float(toplam_alacak))
    alacak_t.number_format = '#,##0.00'
    alacak_t.alignment = Alignment(horizontal="right")

    # Sutun genislikleri
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    wb.save(dosya_yolu)
    print(f"  Olusturuldu: {dosya_yolu}")


def xlsx_plain_olustur(dosya_yolu, hesaplar):
    """Duz (formatsiz) Excel mizan dosyasi olustur - parser uyumlu."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mizan"

    # Header (satir 1)
    headers = ["Hesap Kodu", "Hesap Adi", "Borc Bakiye", "Alacak Bakiye"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Veri satirlari
    for idx, (hesap_kodu, hesap_adi, borc, alacak) in enumerate(hesaplar, 2):
        ws.cell(row=idx, column=1, value=hesap_kodu)
        ws.cell(row=idx, column=2, value=hesap_adi)
        ws.cell(row=idx, column=3, value=float(d(borc)))
        ws.cell(row=idx, column=4, value=float(d(alacak)))

    wb.save(dosya_yolu)
    print(f"  Olusturuldu: {dosya_yolu}")


def csv_olustur(dosya_yolu, hesaplar, delimiter=","):
    """CSV mizan dosyasi olustur"""
    with open(dosya_yolu, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerow(["Hesap Kodu", "Hesap Adi", "Borc Bakiye", "Alacak Bakiye"])
        for hesap_kodu, hesap_adi, borc, alacak in hesaplar:
            writer.writerow([hesap_kodu, hesap_adi, f"{borc:.2f}", f"{alacak:.2f}"])
    print(f"  Olusturuldu: {dosya_yolu}")


def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    print("\nTFRS Converter - Mock Mizan Verisi Uretici")
    print("=" * 60)

    # --- BASIT MIZAN ---
    basit_denge, _, _ = hesap_denge_kontrol(BASIT_HESAPLAR, "Basit Mizan")
    basit_gelir = gelir_tablosu_kontrol(BASIT_HESAPLAR, "Basit Mizan")

    if not basit_denge:
        print("\n  HATA: Basit mizan dengesiz! Dosya olusturulmayacak.")
        return
    if not basit_gelir:
        print("\n  HATA: Basit mizan gelir tablosu 590 ile uyumsuz!")
        return

    basit_xlsx = os.path.join(base_dir, "mizan_basit.xlsx")
    xlsx_olustur(basit_xlsx, BASIT_HESAPLAR, "ABC Ticaret A.S.", "01.01.2024 - 31.12.2024")

    basit_xlsx_plain = os.path.join(base_dir, "mizan_basit_plain.xlsx")
    xlsx_plain_olustur(basit_xlsx_plain, BASIT_HESAPLAR)

    basit_csv = os.path.join(base_dir, "mizan_basit.csv")
    csv_olustur(basit_csv, BASIT_HESAPLAR)

    # --- KAPSAMLI MIZAN ---
    kapsamli_denge, _, _ = hesap_denge_kontrol(KAPSAMLI_HESAPLAR, "Kapsamli Mizan")
    kapsamli_gelir = gelir_tablosu_kontrol(KAPSAMLI_HESAPLAR, "Kapsamli Mizan")

    if not kapsamli_denge:
        print("\n  HATA: Kapsamli mizan dengesiz! Dosya olusturulmayacak.")
        return
    if not kapsamli_gelir:
        print("\n  HATA: Kapsamli mizan gelir tablosu 590 ile uyumsuz!")
        return

    kapsamli_xlsx = os.path.join(base_dir, "mizan_kapsamli.xlsx")
    xlsx_olustur(
        kapsamli_xlsx, KAPSAMLI_HESAPLAR,
        "XYZ Sanayi ve Ticaret A.S.", "01.01.2024 - 31.12.2024"
    )

    kapsamli_xlsx_plain = os.path.join(base_dir, "mizan_kapsamli_plain.xlsx")
    xlsx_plain_olustur(kapsamli_xlsx_plain, KAPSAMLI_HESAPLAR)

    print("\n" + "=" * 60)
    print("  Tum dosyalar basariyla olusturuldu!")
    print("=" * 60)


if __name__ == "__main__":
    main()
